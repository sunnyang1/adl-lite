"""Excel importer — reads Excel files and returns structured data."""

from __future__ import annotations

from typing import Any


class ExcelImporter:
    """Imports and analyzes data from Excel (.xlsx, .xls) files."""

    @staticmethod
    def import_excel(file_path: str, options: dict | None = None) -> dict[str, Any]:
        """
        Read an Excel file and return structured data per sheet.

        Args:
            file_path: Absolute path to the Excel file.
            options: Optional dict with:
                - sheets: list[str] — specific sheets to import (default: all)
                - max_rows_per_sheet: int (default 0 = unlimited)
                - has_header: bool (default True)

        Returns:
            Dict with keys: sheets (list of {name, headers, data, row_count}).
        """
        opts = options or {}
        target_sheets = opts.get("sheets", [])
        max_rows = opts.get("max_rows_per_sheet", 0)
        has_header = opts.get("has_header", True)

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = target_sheets if target_sheets else wb.sheetnames

            sheets_result: list[dict[str, Any]] = []

            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    sheets_result.append(
                        {
                            "name": sheet_name,
                            "error": "Sheet not found.",
                            "headers": [],
                            "data": [],
                            "row_count": 0,
                        }
                    )
                    continue

                ws = wb[sheet_name]
                raw_rows: list[tuple] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if max_rows > 0 and i >= max_rows + 1:
                        break
                    raw_rows.append(row)

                if not raw_rows:
                    sheets_result.append(
                        {
                            "name": sheet_name,
                            "headers": [],
                            "data": [],
                            "row_count": 0,
                        }
                    )
                    continue

                if has_header:
                    headers = [
                        str(cell) if cell is not None else f"col_{j}"
                        for j, cell in enumerate(raw_rows[0])
                    ]
                    data_rows = raw_rows[1:]
                else:
                    headers = [f"col_{i}" for i in range(len(raw_rows[0]))]
                    data_rows = raw_rows

                data: list[dict[str, str]] = []
                for row in data_rows:
                    row_dict: dict[str, str] = {}
                    for j, header in enumerate(headers):
                        val = row[j] if j < len(row) else ""
                        row_dict[header] = str(val) if val is not None else ""
                    data.append(row_dict)

                # Infer column types
                columns_info: list[dict] = []
                if data:
                    from adl_lite.fde.rule_engine import RuleEngine

                    classification = RuleEngine.classify_data(data)
                    columns_info = classification.get("columns", [])

                sheets_result.append(
                    {
                        "name": sheet_name,
                        "headers": headers,
                        "data": data,
                        "row_count": len(data),
                        "columns": columns_info,
                    }
                )

            wb.close()
            return {"sheets": sheets_result}

        except ImportError:
            return {
                "sheets": [
                    {
                        "name": "error",
                        "error": "openpyxl is required for Excel import.",
                        "headers": [],
                        "data": [],
                        "row_count": 0,
                    }
                ],
            }
